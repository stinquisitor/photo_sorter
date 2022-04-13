import asyncio
import os
import re
from datetime import datetime

from pathlib import Path
from openpyxl import load_workbook

import constants
from tasks import TaskCreator


class BaseSorter:
    def __init__(self, table: str, unsorted: str, outdir: str, logger, loop, summary, extension: str,
                 retush_mode: bool, sub_dir: bool):
        self._summary = summary
        self._task_creator = TaskCreator(summary, extension, Path(unsorted), sub_dir)
        self._loop = loop
        self._logger = logger
        self._table = Path(table)
        self._unsorted = Path(unsorted)
        self._outdir = Path(outdir)
        self._outdir.mkdir(exist_ok=True)
        self._retush_mode = retush_mode

    def sort(self):
        raise NotImplementedError

    def _run_tasks(self):
        concurrency = 5
        tasks = self._task_creator.tasks
        while tasks:
            current_tasks = tasks[:concurrency]
            tasks = tasks[concurrency:]
            current_tasks = list(map(lambda task: self._loop.create_task(task.copy_file()),
                                     current_tasks))  # [loop.create_task(foo()), ioloop.create_task(bar())]
            current_tasks = asyncio.wait(current_tasks)
            self._loop.run_until_complete(current_tasks)

    def _write_summary_report(self):
        all_files_num = len([f for f in os.listdir(str(self._unsorted))
                             if os.path.isfile(os.path.join(self._unsorted, f))])
        handled_files_num = len(self._summary.unique_files)
        self._logger.info(f'Обработано файлов {all_files_num}, отобрано - {handled_files_num}.')
        missed_files_num = len(self._summary.miss_files)
        if missed_files_num:
            self._logger.info(f'!!Внимание!!  - Отсутствуют файлы - {missed_files_num} шт.')
            for num, _, metadata in self._summary.miss_files:
                self._logger.info(f'{metadata} - {num}')

    def get_size_to_folder(self, settings):
        size_to_folder = {}
        for pos, size in enumerate(settings['размер']):
            if size is not None:
                # приведём к верхнему регистру и заменим русскую А на английскую.
                size_to_folder[str(size).upper().replace('А', 'A')] = settings['папка для складывания'][pos]
        return size_to_folder

    # получаем настройки с листа с настройками.
    # лист с настройками - лист с наименованием settings
    # wb -> объект класса workbook с которым работаем
    def _get_settings(self, wb):
        # настроечного листа нет в книге. Пишем ошибку и прекращаем работу.
        if constants.settings_sheetname not in wb.sheetnames:
            self._logger.error('Отсутствует лист с настройками. Название листа с настройками должно быть: ' +
                               constants.settings_sheetname)
            raise Exception
        for ws in wb.worksheets:
            if ws.title == constants.settings_sheetname:
                ws_ = ws
                break
        # заполняем настроечные данные
        head = None
        settings = {}
        # ищем первую строку в которой есть данные в первом столбце.
        for row in ws_.values:
            values = tuple(row)
            # первый столбец не пустой - проверяем что за настройка в нём - нужная ли она для программы.
            # если нет - пропускаем. Список настроек в константах.
            if str(values[0]).lower() in constants.settings_list_reg:
                # просто сохраняем картежем. Так мы сможем напрямую замапить данные при обработке.
                # Для этого пересохраняем кортеж со сдвигом.
                settings[str(values[0]).lower()] = values
        if len(settings) == 0:
            self._logger.error('Отсутствуют данные в листе с настройками.')
            raise Exception
        return settings

    # для сложного формата: 1234(2шт)-A4
    # сначала разбиваем по запятым и строкам
    # потом каждый элемент разбиваем по - в левой части получаем номер фото и количество, в правой - размер
    # (хихи, вот и пригодился размер фото)
    @classmethod
    def _get_num_count_complex(cls, s: str):
        # приведём все А к одному виду (английскому)
        s = s.replace('А', 'A')
        # 1. делаем разбивку по символам новой строки или ,
        all_list = []
        list = s.split('\n')
        for l in list:
            all_list.extend(l.split(','))
        # теперь каждый раскладываем по -
        # тут будет всех по одному.
        for l in all_list:
            a = l.split('-')

            multiple_num = re.findall('(\d+)\s*\(?(\d+)\s*шт', a[0])
            if len(multiple_num) > 0:
                yield multiple_num[0][0], multiple_num[0][1], a[1].strip()
            else:
                all_num = re.findall('(\d+)', a[0])
                if len(all_num) > 0:
                    yield all_num[0], 1, a[1].strip()

    # для простого формата. поддерживаются следующие виды:
    # 1. просто число, например 1234
    # 2. число + указание количества, например 1234 (2 шт)
    # 3. диапазон, например 1234-1244
    @classmethod
    def _get_num_count(cls, s: str):
        try:
            all_num = set(re.findall('(\d+)', s))
        except TypeError:
            print(s)
            return
        multiple_num = re.findall('(\d+)\s*\(?(\d+)\s*шт', s)
        range_num = re.findall('(\d+)[ ]*-[ ]*(\d+).*', s)
        s1 = set(map(lambda x: x[0], multiple_num))
        s2 = set(map(lambda x: x[1], multiple_num))
        s3 = set(map(lambda x: x[0], range_num))
        s4 = set(map(lambda x: x[1], range_num))
        all_num = all_num - s1 - s2 - s3 - s4

        for val in multiple_num:
            yield val
        for val_st, val_fl in range_num:
            for i in range(int(val_st), int(val_fl) + 1):
                yield (str(i), 1)
        for val in all_num:
            yield (val, 1)

    # для ретуширования нам не нужно число копий. оставляем оригинальные названия.
    # функция для складывания в одну папку.
    def _get_by_formats_new_retush(self, val, outdir, third_gift, complex):
        if complex:
            numbers = self._get_num_count_complex(val)
            for num, count, size in numbers:
                self._task_creator.add_task(self._unsorted, num, outdir, self._logger,
                                            copies=int(1), metadata='')
        else:
            numbers = self._get_num_count(val)
            for num, count in numbers:
                self._task_creator.add_task(self._unsorted, num, outdir, self._logger,
                                            copies=int(1), metadata='')


class PrintingSorter(BaseSorter):
    def __init__(self, table: str, unsorted: str, outdir: str, logger, loop, summary, extension: str,
                 retush_mode: bool, subdir_include: bool):
        super().__init__(table, unsorted, outdir, logger, loop, summary, extension, retush_mode, subdir_include)

    def sort(self):
        self._logger.clear()
        wb = load_workbook(self._table, data_only=True)
        # настройки
        try:
            settings = self._get_settings(wb)
        except Exception as err:
            return
        size_to_folder = self.get_size_to_folder(settings)
        # лист с данными

        ws = wb.worksheets[0]

        table_head = None
        # идём построчно
        for row in ws.values:
            values = tuple(row)
            # увидели в столбце B символ № - значит нашли "голову"
            if '№' == values[1]:
                table_head = values
                # когда нашли голову - идём на следующую строку и сохраняем значения "головы".
                continue
            # ищем теперь значащие строки.
            # в значащих строках в values[1] будут числа.
            if table_head:
                # если values[1] можно привести к целому числу - значит строка со значениями.
                try:
                    num = int(values[1])
                    # заодно сразу фио сохраним (удалим переводы строк ненужные)
                    name = values[2].replace('\n', '')
                    # заодно сразу проверим что ФИО не пустое
                except (TypeError, ValueError, AttributeError):
                    continue

                # будем итерироваться по длине строки.
                # начнём с 3-го столбца
                for l in range(2, len(table_head)):
                    # если значений нет - сразу дальше.
                    if values[l] is None:
                        continue
                    # смотрим по настройкам и текущему варианту распределения (для ретуши в одну папку, или нет)
                    if settings['распределять'][l] == 'ДА':
                        # проверяем, разложение для ретуши или нет.
                        # в режиме ретуши и раскладываем
                        if self._retush_mode and settings['раскладывать для ретуши'][l] == 'ДА':
                            # в данном режиме всё кидаем просто в outdir
                            val = str(values[l])
                            try:
                                self._get_by_formats_new_retush(val, self._outdir,
                                                            str(settings['третье в подарок?'][l]).upper() == 'ДА',
                                                            str(settings['сложный формат'][l]).upper() == 'ДА')
                            except Exception as err:
                                self._logger.error('Ошибка при обработке строки (ретушь): '+str(values[1]))
                                self._logger.error('Столбец: ' + str(l))
                                self._logger.error('Значение: ' + val)
                                self._logger.error('Текст ошибки: ' + str(err))
                                self._logger.error('-------------------------------------')
                                continue
                        # не в режиме ретуши, раскладываем по папкам
                        elif not self._retush_mode:
                            val = str(values[l])
                            try:
                                outdir = self._outdir / settings['папка для складывания'][l]
                                outdir = Path(str(outdir).replace('_name_', name))
                            except Exception as err:
                                self._logger.error('Ошибка при определении пути для складывания: ')
                                self._logger.error('Столбец: ' + str(l))
                                self._logger.error('Значение: ' + settings['папка для складывания'][l])
                                self._logger.error('Текст ошибки: ' + str(err))
                                continue
                            if str(settings['сложный формат'][l]).upper() == 'ДА':
                                try:
                                    self._get_by_formats_new_complex(val, outdir,
                                                                 str(settings['третье в подарок?'][l]).upper() == 'ДА',
                                                                 size_to_folder)
                                except Exception as err:
                                    self._logger.error('Ошибка при обработке строки (сложный формат): ' + str(values[1]))
                                    self._logger.error('Столбец: ' + str(l))
                                    self._logger.error('Значение: ' + val)
                                    self._logger.error('Текст ошибки: ' + str(err))
                                    self._logger.error('-------------------------------------')
                                    continue

                            else:
                                try:
                                    self._get_by_formats_new(val, outdir,
                                                         str(settings['третье в подарок?'][l]).upper() == 'ДА')
                                except Exception as err:
                                    self._logger.error('Ошибка при обработке строки (простой формат): ' + str(values[1]))
                                    self._logger.error('Столбец: ' + str(l))
                                    self._logger.error('Значение: ' + val)
                                    self._logger.error('Текст ошибки: ' + str(err))
                                    self._logger.error('-------------------------------------')
                                    continue

                        else:
                            continue
        self._run_tasks()
        # self._rename_vinjetkas()
        self._summary.show()
        self._write_summary_report()
        self._logger.info('Сортировка окончена')

    # обрабатываем сложный формат
    def _get_by_formats_new_complex(self, val, outdir, third_gift, size_to_folder):
        numbers = self._get_num_count_complex(val)
        for num, count, size in numbers:
            cc = int(count)
            out_dir = Path(str(outdir).replace('_size_', size_to_folder[size]))
            if third_gift:
                if int(count) >= 2:
                    c = int(count) / 2
                    cc = int(count) + int(c)
            self._task_creator.add_task(self._unsorted, num, out_dir, self._logger,
                                        copies=int(cc), metadata='')

    # обрабатываем обычный
    def _get_by_formats_new(self, val, outdir, third_gift):
        numbers = self._get_num_count(val)
        for num, count in numbers:
            cc = int(count)
            if third_gift:
                if int(count) >= 2:
                    c = int(count) / 2
                    cc = int(count) + int(c)
            self._task_creator.add_task(self._unsorted, num, outdir, self._logger,
                                        copies=int(cc), metadata='')

    ################ OLD CODE
    # потом удалю.
    def sort_old(self):
        self._logger.clear()
        wb = load_workbook(self._table)
        ws = wb.worksheets[0]
        table_head = None
        table_sub_head = None
        # идём построчно
        for row in ws.values:
            values = tuple(row)
            # увидели в столбце B символ № - значит нашли "голову"
            if '№' == values[1]:
                table_head = values
                # следующая итрерация должна получить вторую строку заголовка!
                table_sub_head = None
                # когда нашли голову - идём на следующую строку и сохраняем значения "головы".
                continue
            # если сюда пришли и "голова" есть - включаем второй этап. Нам нужно найти стоблбцы значащие
            # и распределить их по группам. Есть группы стандартные, а есть для фото в эл. виде
            if table_head and table_sub_head is None:
                table_sub_head = values
                continue

            # нашли строку со значениями под заголовком
            if table_head and table_sub_head:
                # быстренько проверим, если строка - пустая целиком (все None) - значит это строка между значениями.
                # пропускаем тогда и установим всё в None.
                # криво
                flag = False
                for l in range(len(table_head)):
                    if values[l]:
                        flag = True
                        break
                if not flag:
                    table_head = None
                    table_sub_head = None
                    continue
                # будем итерироваться по длинне строки.
                # поскольку нужно следить за значениями сразу в трёх строках
                # начнём с 3-го столбца
                for l in range(2, len(table_head)):
                    # если значений нет - сразу дальше.
                    if values[l] is None:
                        continue
                    # первое - сначала только те столбцы где во второй строке стоит №
                    if table_sub_head[l] is not None and table_sub_head[l].rstrip() == '№':
                        head = table_head[l]
                        val = values[l]
                        self._get_by_formats_single(head, val)
                    # ищем стоблец с подарками
                    if table_head[l] is not None and table_head[l].lower() == 'Номер и размер фото в подарок'.lower():
                        self._get_gift_new(values[l])
                    # if table_head[l].lower() == 'Номер и размер фото в подарок'.lower():
                # обработали все столбцы с фактически
                # head_values = [x.split()[0] for x in table_head[6:-12:2] if x is not None]
                # head_second_values = [str(x) for x in table_sub_head[6:-12:2]]
                # [str(x) for x in values[6:-12:2]]
                # self._get_by_formats(['="Общая фотография'],
                #                     [str(x) for x in [values[4]]])
                # self._get_by_formats(['="Виньетки'],
                #                     [str(x) for x in [values[4]]])
                # self._get_common(str(values[5]))
                # self._get_gift(values[-6:-3])
                # self._get_gift(values[-9:-6])

        self._run_tasks()
        # self._rename_vinjetkas()
        self._summary.show()
        self._write_summary_report()
        self._logger.info('Сортировка окончена')

    def _get_by_formats(self, head, values):
        for i in range(0, len(values)):
            if values[i] is None:
                continue
            numbers = self._get_num_count(values[i])
            for num, count in numbers:
                out_dir = self._outdir / head[i].split('"')[1].rstrip()
                self._task_creator.add_task(self._unsorted, num, out_dir, self._logger,
                                            copies=int(count), metadata='')

    def _get_by_formats_single(self, head, values):
        numbers = self._get_num_count(values)
        for num, count in numbers:
            out_dir = self._outdir / head.split('"')[1].rstrip()
            self._task_creator.add_task(self._unsorted, num, out_dir, self._logger,
                                        copies=int(count), metadata='')

    def _get_gift(self, value):
        num, size, count = value
        if not num or not size or not count:
            return

        if not isinstance(count, int) or not isinstance(num, int):
            return
        out_dir = f'="{size.upper().strip(" ")}'
        for i in range(0, count):
            self._get_by_formats([out_dir], [str(num)])

    # новый способ записи подарков - записывается номер фото и размер, через запятую.
    def _get_gift_new(self, value: str):
        # предполагаем следующий формат: номер_фото размер (через пробел)
        for gift in value.split(','):
            try:
                num, size = gift.lstrip().rstrip().split(' ')
            except Exception as err:
                self._logger.info('Ошибка формата в столбце с подарками:')
                self._logger.info(value)
                continue
            if not num or not size:
                continue
            out_dir = f'="{size.upper().strip(" ")}'
            # for i in range(0, count):
            self._get_by_formats([out_dir], [str(num)])

    def _rename_vinjetkas(self):
        dir_name = 'Виньетки'
        summary_result = self._summary.get(dir_name)
        if summary_result is None:
            return
        dir_name, count = summary_result
        dir_name.rename(dir_name.parent / f'{dir_name.name}_{count}')

    def _get_common(self, value):
        if value is None:
            return
        try:
            srange = value.replace(" ", "").split("-")
            irange = list(map(lambda x: int(x), srange))
            if len(irange) == 2:
                irange[1] += 1
            srange = list(map(lambda x: str(x), range(*irange)))
            self._get_by_formats(['="Электронный вид'] * len(srange),
                                 srange)
        except Exception as te:
            pass


# основное отличие - разделение по видам альбомов. макси и стандарт.
# на данный момент - в альбомах вся информация идёт подряд, без пропусков и без лишних заголовков.
class AlbumSorter(BaseSorter):
    def __init__(self, table: str, unsorted: str, outdir: str, logger, loop, summary, extension: str,
                 retush_mode: bool, sub_dir: bool):
        super().__init__(table, unsorted, outdir, logger, loop, summary, extension, retush_mode, sub_dir)

    def sort(self):
        self._logger.clear()
        wb = load_workbook(self._table)
        # настройки
        try:
            settings = self._get_settings(wb)
        except Exception as err:
            return
        size_to_folder = self.get_size_to_folder(settings)
        # лист с данными

        ws = wb.worksheets[0]

        table_head = None
        # идём построчно
        for row in ws.values:
            values = tuple(row)
            # увидели в столбце B символ № - значит нашли "голову"
            if '№' == values[1]:
                table_head = values
                # когда нашли голову - идём на следующую строку и сохраняем значения "головы".
                continue

            # нашли строку со значениями под заголовком
            if table_head:
                # быстренько проверим, если строка - пустая (достаточно первых значащих
                # пропускаем тогда и установим всё в None.
                # криво
                flag = False
                for l in range(len(table_head)):
                    if values[l]:
                        flag = True
                        break
                if not flag:
                    table_head = None
                    continue
                # Будем итерироваться по длине строки.
                # Поскольку нужно следить за значениями сразу в трёх строках
                # начнём с 3-го столбца
                for l in range(2, len(table_head)):
                    # если значений нет - сразу дальше.
                    if values[l] is None:
                        continue
                    # смотрим по настройкам и текущему варианту распределения (для ретуши в одну папку, или нет)
                    if settings['распределять'][l] == 'ДА':
                        # проверяем, разложение для ретуши или нет.
                        # в режиме ретуши и раскладываем
                        if self._retush_mode and settings['раскладывать для ретуши'][l] == 'ДА':
                            # в данном режиме всё кидаем просто в outdir
                            val = values[l]
                            self._get_by_formats_new_retush(val, self._outdir,
                                                            str(settings['третье в подарок?'][l]).upper() == 'ДА',
                                                            str(settings['сложный формат'][l]).upper() == 'ДА')
                        # не в режиме ретуши, раскладываем по папкам
                        elif not self._retush_mode:
                            val = values[l]
                            outdir = self._outdir / settings['папка для складывания'][l]
                            if str(settings['сложный формат'][l]).upper() == 'ДА':
                                self._get_by_formats_new_complex(val, outdir,
                                                                 str(settings['третье в подарок?'][l]).upper() == 'ДА',
                                                                 size_to_folder)
                            else:
                                self._get_by_formats_new(val, outdir,
                                                         str(settings['третье в подарок?'][l]).upper() == 'ДА')
                        else:
                            continue
        self._run_tasks()
        # self._rename_vinjetkas()
        self._summary.show()
        self._write_summary_report()
        self._logger.info('Сортировка окончена')

    ###########################OLD CODE
    def sort_old(self):
        self._logger.clear()
        wb = load_workbook(self._table)
        ws = wb.worksheets[0]
        table_head = None
        kid_counter = 0

        out_dir = self._outdir / '2 сортировка - папка для каждого ребенка, обложка, виньетка, групповые'
        out_dir.mkdir(exist_ok=True)

        out_all_photo_dir = self._outdir / '1 сортировка - все фотки в одну папку'
        out_all_photo_dir.mkdir(exist_ok=True)

        for row in ws.values:
            table_values = tuple(row)
            self._sort_common(table_values,
                              out_dir / '0.Групповые фото',
                              out_all_photo_dir)
            if '№' == table_values[0]:
                table_head = table_values
                continue
            if table_head:
                kid_counter += 1
                if not table_values[1]:
                    continue

                subdir_counter = 0
                for i in range(3, len(table_values)):
                    value = table_values[i]
                    if not value:
                        continue
                    if not table_head[i]:
                        continue
                    if not isinstance(table_head[i], str):
                        continue

                    from_head = table_head[i][:30].strip()

                    if from_head.lower() == 'пожелания':
                        continue

                    metadata = ''
                    subdir_counter += 1
                    if i == 3 or i == 4:
                        metadata = f'{kid_counter}.{table_values[1]} - 0.{from_head}'
                        out_photo_dir = out_dir / f'0.{from_head}'
                    else:
                        dir_name = f'{kid_counter}.{table_values[1]}'
                        Path(out_dir / dir_name).mkdir(exist_ok=True)
                        out_photo_dir = out_dir / dir_name / f'{subdir_counter - 2}.{from_head}'
                        metadata = f'{dir_name} - {subdir_counter - 2}.{from_head}'
                    out_photo_dir.mkdir(exist_ok=True)
                    all_num = set(re.findall('(\d+)', str(value)))
                    for num in all_num:
                        self._task_creator.add_task(self._unsorted, str(num), out_photo_dir, self._logger,
                                                    copies=0, metadata=metadata)
                        self._task_creator.add_task(self._unsorted, str(num), out_all_photo_dir, self._logger,
                                                    copies=0, summary=False)

        self._run_tasks()
        self._write_summary_report()
        self._logger.info('Сортировка окончена')

    def _sort_common(self, values, out_photo_dir, out_all_photo_dir):
        if not isinstance(values[0], str):
            return
        if not values[0].lower().startswith('общие групповые фотографии в альбом'):
            return
        for value in values:
            if not value:
                continue
            all_num = set(re.findall('(\d+)', str(value)))
            out_photo_dir.mkdir(exist_ok=True)
            for num in all_num:
                self._task_creator.add_task(self._unsorted, str(num), out_photo_dir, self._logger, copies=0)
                self._task_creator.add_task(self._unsorted, str(num), out_all_photo_dir, self._logger,
                                            copies=0, summary=False)
